from selenium import webdriver
import time
import csv
import requests
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import chromedriver_autoinstaller

import requests
import re
import traceback
from bs4 import BeautifulSoup

headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36'
}

from openpyxl import load_workbook, Workbook

sheet_name = 'data.xlsx'

# Write Headline and create a new excel sheet
def xl_sheet_headlines(sheet_name=sheet_name):
    wb = Workbook()
    ws = wb.active
    headlines = ['url', 'name', 'address', 'website', 'phone', 'category', 'email']
    ws.append(headlines)
    wb.save(sheet_name)
xl_sheet_headlines()
    
# Write Data On existing sheet
def xl_write(data_write, sheet_name=sheet_name):
    wb = load_workbook(sheet_name)
    work_sheet = wb.active # Get active sheet
    work_sheet.append(data_write)
    wb.save(sheet_name)

def driver_define():
    print('Chromedriver Installing')
    driver_path = chromedriver_autoinstaller.install()
    
    print('Chrome Browser Opening')
    options = Options()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    s = Service(driver_path)
    driver = webdriver.Chrome(service=s, options =options)
    return driver

# Email Get
def get_email(url):
    domain = url.split('//')[-1].replace('www.', '').split('/')[0]
    url_gen = f'http://www.skymem.info/srch?q={domain}'
    response = requests.get(url_gen, headers=headers)
    soup = BeautifulSoup(response.text, 'lxml')
    email_list = re.findall(r"href=\"\/srch\?q=(.*?@.*)\">", str(soup))
    email = [line for line in email_list if domain in line][0]
    
    return email

driver = driver_define()
urls_filename = 'urls.txt'
urls = [line.strip('\n') for line in open(urls_filename).readlines()]

for url in urls:

    driver.get(url)

    print('--------------------------')

    try:
        name = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//h1'))).text
    except:
        name = 'N/A'

    time.sleep(1)

    try:
        address = driver.find_element(By.XPATH, '//button[@data-item-id="address"]').text
    except:
        address = 'N/A'

    try:
        website = driver.find_element(By.CSS_SELECTOR, 'a[aria-label^="Website:"]').get_attribute('href')
    except:
        website = 'N/A'

    try:
        phone = driver.find_element(By.CSS_SELECTOR, 'button[aria-label*="Phone:"]').text
    except:
        phone = 'N/A'

    try:
        category = driver.find_element(By.CSS_SELECTOR, '[jsaction="pane.rating.category"]').text
    except:
        category = 'N/A'

    email = 'N/A'
    try:
        if len(website) > 3:
            email = get_email(website)
    except:
        email = 'N/A'

    print(f"name : {name}")
    print(f"address : {address}")
    print(f"website:", website)
    print(f"phone:", phone)
    print("category:", category)
    print("email:", email)

    write_data = [url, name, address, website, phone, category, email]

    xl_write(write_data)
